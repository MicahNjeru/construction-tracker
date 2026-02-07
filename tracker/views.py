from collections import defaultdict
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, F
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, FileResponse, Http404, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from .forms import *
from labor.models import LaborEntry
import os

# Create your views here.


# ==================== User Profile ====================

@login_required
def user_profile(request):
    """View and edit user profile."""
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('user_profile')
    else:
        form = UserProfileForm(instance=profile)
    
    context = {
        'form': form,
        'profile': profile,
    }
    return render(request, 'tracker/user_profile.html', context)


@login_required
def dashboard(request):
    """Dashboard with overview and analytics - Phase 2: Enhanced."""
    projects = Project.objects.filter(created_by=request.user)
    
    # Summary statistics
    total_projects = projects.count()
    active_projects = projects.filter(status='in_progress').count()
    completed_projects = projects.filter(status='completed').count()
    total_budget = projects.aggregate(Sum('budget'))['budget__sum'] or 0
    total_spent = sum(p.total_spent for p in projects)
    
    # Calculate total material cost
    total_material_cost = MaterialEntry.objects.filter(
        project__created_by=request.user
    ).aggregate(total=Sum('cost'))['total'] or Decimal('0.00')

    # Calculate total labor cost
    total_labor_cost = LaborEntry.objects.filter(
        project__created_by=request.user
    ).aggregate(
        total=Sum(F('number_of_workers') * F('rate_per_worker_per_day'))
    )['total'] or Decimal('0.00')

    # Recent materials entries
    recent_materials = MaterialEntry.objects.filter(
        project__created_by=request.user
    ).order_by('-created_at')[:5]

    # Recent labor entries
    recent_labor = LaborEntry.objects.filter(
        project__created_by=request.user
    ).select_related('project', 'category').order_by('-work_date', '-created_at')[:10]
    
    # Material type breakdown by category
    material_type_stats = MaterialEntry.objects.filter(
        project__created_by=request.user
    ).values('category__name').annotate(
        total_cost=Sum('cost'),
        count=Count('id')
    ).order_by('-total_cost')[:5]

    # Labor breakdown by category
    labor_breakdown = LaborEntry.objects.filter(
        project__created_by=request.user
    ).values('category__name').annotate(
        total_cost=Sum(F('number_of_workers') * F('rate_per_worker_per_day'))
    ).order_by('-total_cost')[:10]
    
    # Monthly spending
    # ---- Materials grouped by month ----
    materials_monthly = (
        MaterialEntry.objects
        .filter(project__created_by=request.user)
        .annotate(month=TruncMonth('purchase_date'))
        .values('month')
        .annotate(material_cost=Sum('cost'))
    )

    # ---- Labor grouped by month ----
    labor_monthly = (
        LaborEntry.objects
        .filter(project__created_by=request.user)
        .annotate(month=TruncMonth('work_date'))
        .values('month')
        .annotate(
            labor_cost=Sum(F('number_of_workers') * F('rate_per_worker_per_day'))
        )
    )

    # ---- Merge into single timeline ----
    monthly_map = defaultdict(lambda: {
        'material_cost': 0,
        'labor_cost': 0
    })

    for row in materials_monthly:
        monthly_map[row['month']]['material_cost'] = row['material_cost']

    for row in labor_monthly:
        monthly_map[row['month']]['labor_cost'] = row['labor_cost']

    monthly_spending = [
        {
            'month': month,
            'material_cost': data['material_cost'],
            'labor_cost': data['labor_cost'],
        }
        for month, data in sorted(monthly_map.items())
    ]
    
    context = {
        'total_projects': total_projects,
        'active_projects': active_projects,
        'completed_projects': completed_projects,
        'total_budget': total_budget,
        'total_spent': total_spent,
        'total_material_cost': total_material_cost,
        'total_labor_cost': total_labor_cost,
        'recent_materials': recent_materials,
        'labor_breakdown': labor_breakdown,
        'recent_labor': recent_labor,
        'projects': projects[:5],
        'material_type_stats': material_type_stats,
        'monthly_spending': monthly_spending,
    }
    return render(request, 'tracker/dashboard.html', context)


@login_required
def project_list(request):
    """Display list of all projects with search."""
    projects = Project.objects.filter(created_by=request.user)
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        projects = projects.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(location__icontains=search_query)
        )
    
    # Status filter
    status_filter = request.GET.get('status', '')
    if status_filter:
        projects = projects.filter(status=status_filter)
    
    # Calculate summary statistics
    total_projects = projects.count()
    active_projects = projects.filter(status='in_progress').count()
    completed_projects = projects.filter(status='completed').count()
    total_budget = projects.aggregate(Sum('budget'))['budget__sum'] or 0
    
    context = {
        'projects': projects,
        'total_projects': total_projects,
        'active_projects': active_projects,
        'completed_projects': completed_projects,
        'total_budget': total_budget,
        'search_query': search_query,
        'status_filter': status_filter,
        'status_choices': Project.STATUS_CHOICES,
    }
    return render(request, 'tracker/project_list.html', context)


@login_required
def project_detail(request, pk):
    """Display project details and materials with filtering."""
    project = get_object_or_404(Project, pk=pk)
    materials = project.material_entries.all()

    # Labor entries
    labor_entries = project.labor_entries.all()
    
    # Search materials
    search_query = request.GET.get('search', '')
    if search_query:
        materials = materials.filter(
            Q(description__icontains=search_query) |
            Q(supplier__icontains=search_query) |
            Q(notes__icontains=search_query)
        )
    
    # Filter by material type
    type_filter = request.GET.get('type', '')
    if type_filter:
        materials = materials.filter(category=type_filter)
    
    # Filter by date range
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        materials = materials.filter(purchase_date__gte=date_from)
    if date_to:
        materials = materials.filter(purchase_date__lte=date_to)

    # Material breakdown by category
    material_breakdown = project.material_entries.values('category__name').annotate(
        total=Sum('cost')
    ).order_by('-total')

    # Labor breakdown by category
    labor_breakdown = project.labor_entries.values('category__name').annotate(
        total_cost=Sum(F('number_of_workers') * F('rate_per_worker_per_day'))
    ).order_by('-total_cost')
    
    context = {
        'project': project,
        'materials': materials,
        'labor_entries': labor_entries,
        'material_breakdown': material_breakdown,
        'labor_breakdown': labor_breakdown,
        'search_query': search_query,
        'type_filter': type_filter,
        'date_from': date_from,
        'date_to': date_to,
        'material_categories': MaterialCategory.objects.all(),
    }
    return render(request, 'tracker/project_detail.html', context)


@login_required
def project_create(request):
    """Create a new project."""
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.created_by = request.user
            project.save()
            messages.success(request, f'Project "{project.name}" created successfully!')
            return redirect('project_detail', pk=project.pk)
    else:
        form = ProjectForm()
    
    return render(request, 'tracker/project_form.html', {'form': form, 'title': 'Create Project'})


@login_required
def project_update(request, pk):
    """Update an existing project."""
    project = get_object_or_404(Project, pk=pk)
    
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            messages.success(request, f'Project "{project.name}" updated successfully!')
            return redirect('project_detail', pk=project.pk)
    else:
        form = ProjectForm(instance=project)
    
    return render(request, 'tracker/project_form.html', {
        'form': form, 
        'title': 'Update Project',
        'project': project
    })


@login_required
def project_delete(request, pk):
    """Delete a project."""
    project = get_object_or_404(Project, pk=pk)
    
    if request.method == 'POST':
        project_name = project.name
        project.delete()
        messages.success(request, f'Project "{project_name}" deleted successfully!')
        return redirect('project_list')
    
    return render(request, 'tracker/project_confirm_delete.html', {'project': project})


@login_required
def material_create(request, project_pk):
    """Create a new material entry for a project."""
    project = get_object_or_404(Project, pk=project_pk)
    
    if request.method == 'POST':
        form = MaterialEntryForm(request.POST)
        if form.is_valid():
            material = form.save(commit=False)
            material.project = project
            material.created_by = request.user
            material.save()
            ActivityLog.objects.create(
                project=project,
                user=request.user,
                action='material_added',
                description=f"Added {material.category.name}: {material.description}",
                related_material=material
            )
            check_budget_alerts(project)
            messages.success(request, 'Material entry added successfully!')
            return redirect('project_detail', pk=project.pk)
    else:
        form = MaterialEntryForm()
    
    return render(request, 'tracker/material_form.html', {
        'form': form,
        'project': project,
        'title': 'Add Material'
    })


@login_required
def material_update(request, pk):
    """Update an existing material entry."""
    material = get_object_or_404(MaterialEntry, pk=pk)
    project = material.project
    
    if request.method == 'POST':
        form = MaterialEntryForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            ActivityLog.objects.create(
                project=project,
                user=request.user,
                action='material_updated',
                description=f"Updated {material.category.name}: {material.description}",
                related_material=material
            )
            check_budget_alerts(project)
            messages.success(request, 'Material entry updated successfully!')
            return redirect('project_detail', pk=project.pk)
    else:
        form = MaterialEntryForm(instance=material)
    
    return render(request, 'tracker/material_form.html', {
        'form': form,
        'project': project,
        'title': 'Update Material',
        'material': material
    })


@login_required
def material_delete(request, pk):
    """Delete a material entry."""
    material = get_object_or_404(MaterialEntry, pk=pk)
    project = material.project
    
    if request.method == 'POST':
        ActivityLog.objects.create(
            project=project,
            user=request.user,
            action='material_deleted',
            description=f"Deleted {material.category.name}: {material.description}"
        )
        material.delete()
        messages.success(request, 'Material entry deleted successfully!')
        return redirect('project_detail', pk=project.pk)
    
    return render(request, 'tracker/material_confirm_delete.html', {
        'material': material,
        'project': project
    })


@login_required
def receipt_upload(request, material_pk):
    """Upload a receipt for a material entry - Phase 2: Multiple receipts supported."""
    material = get_object_or_404(MaterialEntry, pk=material_pk)
    
    if request.method == 'POST':
        form = ReceiptUploadForm(request.POST, request.FILES)
        if form.is_valid():
            receipt = form.save(commit=False)
            receipt.material_entry = material
            receipt.save()
            
            messages.success(request, 'Receipt uploaded successfully!')
            return redirect('project_detail', pk=material.project.pk)
    else:
        form = ReceiptUploadForm()
    
    return render(request, 'tracker/receipt_upload.html', {
        'form': form,
        'material': material,
        'project': material.project,
        'existing_receipts': material.receipts.all()
    })


@login_required
def receipt_gallery(request, material_pk):
    """View all receipts for a material entry in gallery format."""
    material = get_object_or_404(MaterialEntry, pk=material_pk)
    receipts = material.receipts.all()
    
    context = {
        'material': material,
        'project': material.project,
        'receipts': receipts,
    }
    return render(request, 'tracker/receipt_gallery.html', context)


@login_required
def receipt_set_primary(request, pk):
    """Set a receipt as primary."""
    receipt = get_object_or_404(Receipt, pk=pk)
    
    # Unmark all other receipts as primary
    receipt.material_entry.receipts.update(is_primary=False)
    
    # Mark this one as primary
    receipt.is_primary = True
    receipt.save()
    
    messages.success(request, 'Primary receipt updated!')
    return redirect('receipt_gallery', material_pk=receipt.material_entry.pk)


@login_required
def receipt_view(request, pk):
    """View or download a receipt."""
    receipt = get_object_or_404(Receipt, pk=pk)
    
    if receipt.file:
        try:
            return FileResponse(
                receipt.file.open('rb'),
                content_type='application/octet-stream',
                as_attachment=False,
                filename=receipt.original_filename
            )
        except Exception as e:
            raise Http404("Receipt file not found.")
    else:
        raise Http404("Receipt file not found.")


@login_required
def receipt_download(request, pk):
    """Download a receipt."""
    receipt = get_object_or_404(Receipt, pk=pk)
    
    if receipt.file:
        try:
            return FileResponse(
                receipt.file.open('rb'),
                content_type='application/octet-stream',
                as_attachment=True,
                filename=receipt.original_filename
            )
        except Exception as e:
            raise Http404("Receipt file not found.")
    else:
        raise Http404("Receipt file not found.")


@login_required
def receipt_delete(request, pk):
    """Delete a receipt."""
    receipt = get_object_or_404(Receipt, pk=pk)
    material = receipt.material_entry
    project = material.project
    
    if request.method == 'POST':
        receipt.delete()
        messages.success(request, 'Receipt deleted successfully!')
        
        # Redirect to gallery if there are more receipts, otherwise to project detail
        if material.receipts.exists():
            return redirect('receipt_gallery', material_pk=material.pk)
        else:
            return redirect('project_detail', pk=project.pk)
    
    return render(request, 'tracker/receipt_confirm_delete.html', {
        'receipt': receipt,
        'material': material,
        'project': project
    })


@login_required
def export_project_excel(request, pk):
    """Export project materials to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'openpyxl is required for Excel export. Install with: pip install openpyxl')
        return redirect('project_detail', pk=pk)
    
    project = get_object_or_404(Project, pk=pk)
    materials = project.material_entries.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = ['Date', 'Type', 'Description', 'Quantity', 'Unit', 'Cost', 'Supplier', 'Has Receipt', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, material in enumerate(materials, 2):
        ws.cell(row=row, column=1, value=material.purchase_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=material.category.name)
        ws.cell(row=row, column=3, value=material.description)
        ws.cell(row=row, column=4, value=float(material.quantity))
        ws.cell(row=row, column=5, value=material.unit.abbreviation)
        ws.cell(row=row, column=6, value=float(material.cost))
        ws.cell(row=row, column=7, value=material.supplier)
        ws.cell(row=row, column=8, value='Yes' if material.has_receipt else 'No')
        ws.cell(row=row, column=9, value=material.notes)
    
    # Summary
    summary_row = len(materials) + 3
    ws.cell(row=summary_row, column=5, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=float(project.total_spent)).font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 30
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{project.name.replace(' ', '_')}_materials_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def export_project_pdf(request, pk):
    """Export project summary to PDF."""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
    except ImportError:
        messages.error(request, 'reportlab is required for PDF export. Install with: pip install reportlab')
        return redirect('project_detail', pk=pk)
    
    project = get_object_or_404(Project, pk=pk)
    materials = project.material_entries.all()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    filename = f"{project.name.replace(' ', '_')}_report_{datetime.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create PDF
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
    )
    elements.append(Paragraph(f"Project Report: {project.name}", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # Project details
    details = [
        ['Location:', project.location or 'N/A'],
        ['Status:', project.get_status_display()],
        ['Start Date:', project.start_date.strftime('%Y-%m-%d')],
        ['Budget:', f'Ksh{project.budget:,.2f}'],
        ['Total Spent:', f'Ksh{project.total_spent:,.2f}'],
        ['Remaining:', f'Ksh{project.remaining_budget:,.2f}'],
    ]
    
    details_table = Table(details, colWidths=[2*inch, 4*inch])
    details_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8e8e8')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(details_table)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Materials table
    elements.append(Paragraph("Materials List", styles['Heading2']))
    elements.append(Spacer(1, 0.1 * inch))
    
    material_data = [['Date', 'Type', 'Description', 'Qty', 'Cost']]
    for material in materials:
        material_data.append([
            material.purchase_date.strftime('%Y-%m-%d'),
            material.category.name,
            material.description[:40],
            f"{material.quantity} {material.unit.abbreviation}",
            f'Ksh{material.cost:,.2f}'
        ])
    
    material_table = Table(material_data, colWidths=[1*inch, 1.2*inch, 2.5*inch, 1*inch, 1*inch])
    material_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
    ]))
    elements.append(material_table)
    
    # Build PDF
    doc.build(elements)
    return response


# ==================== Material Usage Tracking ====================

@login_required
def update_material_usage(request, pk):
    """Update material usage quantity."""
    material = get_object_or_404(MaterialEntry, pk=pk)
    project = material.project
    
    if request.method == 'POST':
        form = MaterialUsageForm(request.POST)
        if form.is_valid():
            new_quantity_used = form.cleaned_data['quantity_used']
            usage_notes = form.cleaned_data.get('notes', '')
            
            # Validate
            if new_quantity_used > material.quantity:
                messages.error(request, 'Quantity used cannot exceed total quantity!')
                return redirect('project_detail', pk=project.pk)
            
            # Update material
            old_quantity_used = material.quantity_used
            material.quantity_used = new_quantity_used
            material.save()
            
            # Log activity
            ActivityLog.objects.create(
                project=project,
                user=request.user,
                action='material_used',
                description=f"Updated usage for {material.description}: {old_quantity_used} → {new_quantity_used} {material.unit.abbreviation}. {usage_notes}",
                related_material=material
            )
            
            messages.success(request, 'Material usage updated successfully!')
            return redirect('project_detail', pk=project.pk)
    else:
        form = MaterialUsageForm(initial={'quantity_used': material.quantity_used})
    
    return render(request, 'tracker/material_usage_form.html', {
        'form': form,
        'material': material,
        'project': project
    })


@login_required
def quick_update_usage(request, pk):
    """Quick AJAX update for material usage."""
    if request.method == 'POST':
        material = get_object_or_404(MaterialEntry, pk=pk)
        quantity_used = request.POST.get('quantity_used')
        
        try:
            quantity_used = float(quantity_used)
            if quantity_used < 0 or quantity_used > float(material.quantity):
                return JsonResponse({'success': False, 'error': 'Invalid quantity'})
            
            material.quantity_used = quantity_used
            material.save()
            
            return JsonResponse({
                'success': True,
                'quantity_remaining': float(material.quantity_remaining),
                'usage_percentage': float(material.usage_percentage)
            })
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid number'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})


# ==================== Project Templates ====================

@login_required
def template_create(request):
    """Create a new project template."""
    if request.method == 'POST':
        form = ProjectTemplateForm(request.POST)
        if form.is_valid():
            template = form.save(commit=False)
            template.created_by = request.user
            template.save()
            messages.success(request, 'Template created successfully!')
            return redirect('template_detail', pk=template.pk)
    else:
        form = ProjectTemplateForm()
    
    return render(request, 'tracker/template_form.html', {
        'form': form,
        'title': 'Create Template'
    })


@login_required
def template_list(request):
    """List all templates available to user."""
    user_templates = ProjectTemplate.objects.filter(created_by=request.user)
    public_templates = ProjectTemplate.objects.filter(is_public=True).exclude(created_by=request.user)
    
    context = {
        'user_templates': user_templates,
        'public_templates': public_templates,
    }
    return render(request, 'tracker/template_list.html', context)


@login_required
def template_detail(request, pk):
    """View template details and materials."""
    template = get_object_or_404(ProjectTemplate, pk=pk)
    materials = template.materials.all()
    
    # Calculate estimated total
    estimated_total = materials.aggregate(Sum('estimated_cost'))['estimated_cost__sum'] or 0
    
    context = {
        'template': template,
        'materials': materials,
        'estimated_total': estimated_total,
    }
    return render(request, 'tracker/template_detail.html', context)


@login_required
def template_update(request, pk):
    """Update an existing project template."""
    template = get_object_or_404(ProjectTemplate, pk=pk)
    
    # Check if user owns the template
    if template.created_by != request.user:
        messages.error(request, 'You do not have permission to edit this template.')
        return redirect('template_list')
    
    if request.method == 'POST':
        form = ProjectTemplateForm(request.POST, instance=template)
        if form.is_valid():
            form.save()
            messages.success(request, f'Template "{template.name}" updated successfully!')
            return redirect('template_detail', pk=template.pk)
    else:
        form = ProjectTemplateForm(instance=template)
    
    return render(request, 'tracker/template_form.html', {
        'form': form,
        'title': 'Update Template',
        'template': template
    })


@login_required
def template_delete(request, pk):
    """Delete a project template."""
    template = get_object_or_404(ProjectTemplate, pk=pk)
    
    # Check if user owns the template
    if template.created_by != request.user:
        messages.error(request, 'You do not have permission to delete this template.')
        return redirect('template_list')
    
    if request.method == 'POST':
        template_name = template.name
        material_count = template.materials.count()
        template.delete()
        messages.success(request, f'Template "{template_name}" and {material_count} materials deleted successfully!')
        return redirect('template_list')
    
    return render(request, 'tracker/template_confirm_delete.html', {
        'template': template,
        'material_count': template.materials.count()
    })


@login_required
def template_add_material(request, template_pk):
    """Add material to template."""
    template = get_object_or_404(ProjectTemplate, pk=template_pk)
    
    if request.method == 'POST':
        form = TemplateMaterialForm(request.POST)
        if form.is_valid():
            material = form.save(commit=False)
            material.template = template
            material.save()
            messages.success(request, 'Material added to template!')
            return redirect('template_detail', pk=template.pk)
    else:
        form = TemplateMaterialForm()
    
    return render(request, 'tracker/template_material_form.html', {
        'form': form,
        'template': template
    })


@login_required
def template_material_update(request, pk, material_pk):
    """Update a material in a template."""
    template = get_object_or_404(ProjectTemplate, pk=pk)
    material = get_object_or_404(TemplateMaterial, pk=material_pk, template=template)
    
    # Check if user owns the template
    if template.created_by != request.user:
        messages.error(request, 'You do not have permission to edit this template.')
        return redirect('template_list')
    
    if request.method == 'POST':
        form = TemplateMaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            messages.success(request, 'Template material updated successfully!')
            return redirect('template_detail', pk=template.pk)
    else:
        form = TemplateMaterialForm(instance=material)
    
    return render(request, 'tracker/template_material_form.html', {
        'form': form,
        'template': template,
        'material': material,
        'title': 'Update Template Material'
    })


@login_required
def template_material_delete(request, pk, material_pk):
    """Delete a material from a template."""
    template = get_object_or_404(ProjectTemplate, pk=pk)
    material = get_object_or_404(TemplateMaterial, pk=material_pk, template=template)
    
    # Check if user owns the template
    if template.created_by != request.user:
        messages.error(request, 'You do not have permission to edit this template.')
        return redirect('template_list')
    
    if request.method == 'POST':
        material_description = material.description
        material.delete()
        messages.success(request, f'Material "{material_description}" removed from template!')
        return redirect('template_detail', pk=template.pk)
    
    return render(request, 'tracker/template_material_confirm_delete.html', {
        'template': template,
        'material': material
    })


@login_required
def create_project_from_template(request):
    """Create a new project from a template."""
    if request.method == 'POST':
        form = CreateProjectFromTemplateForm(request.POST, user=request.user)
        if form.is_valid():
            template = form.cleaned_data['template']
            
            # Create project
            project = Project.objects.create(
                name=form.cleaned_data['name'],
                location=form.cleaned_data['location'],
                budget=form.cleaned_data['budget'],
                start_date=form.cleaned_data['start_date'],
                created_by=request.user,
                created_from_template=template
            )
            
            # Copy materials from template
            for template_material in template.materials.all():
                MaterialEntry.objects.create(
                    project=project,
                    category=template_material.category,
                    description=template_material.description,
                    quantity=template_material.estimated_quantity,
                    unit=template_material.unit,
                    cost=template_material.estimated_cost,
                    purchase_date=form.cleaned_data['start_date'],
                    notes=template_material.notes,
                    created_by=request.user
                )
            
            # Log activity
            ActivityLog.objects.create(
                project=project,
                user=request.user,
                action='project_created',
                description=f"Project created from template '{template.name}' with {template.materials.count()} materials"
            )
            
            messages.success(request, f'Project created from template with {template.materials.count()} materials!')
            return redirect('project_detail', pk=project.pk)
    else:
        form = CreateProjectFromTemplateForm(user=request.user)
    
    return render(request, 'tracker/create_from_template.html', {'form': form})


# ==================== Project Photos ====================

@login_required
def project_photos(request, project_pk):
    """View all photos for a project."""
    project = get_object_or_404(Project, pk=project_pk)
    photos = project.photos.all()
    
    context = {
        'project': project,
        'photos': photos,
    }
    return render(request, 'tracker/project_photos.html', context)


@login_required
def photo_upload(request, project_pk):
    """Upload a photo to project gallery."""
    project = get_object_or_404(Project, pk=project_pk)
    
    if request.method == 'POST':
        form = ProjectPhotoForm(request.POST, request.FILES)
        if form.is_valid():
            photo = form.save(commit=False)
            photo.project = project
            photo.uploaded_by = request.user
            if not photo.taken_date:
                photo.taken_date = timezone.now().date()
            photo.save()
            
            # Log activity
            ActivityLog.objects.create(
                project=project,
                user=request.user,
                action='photo_uploaded',
                description=f"Uploaded photo: {photo.title or 'Untitled'}"
            )
            
            messages.success(request, 'Photo uploaded successfully!')
            return redirect('project_photos', project_pk=project.pk)
    else:
        form = ProjectPhotoForm()
    
    return render(request, 'tracker/photo_upload.html', {
        'form': form,
        'project': project
    })


@login_required
def photo_delete(request, pk):
    """Delete a project photo."""
    photo = get_object_or_404(ProjectPhoto, pk=pk)
    project = photo.project
    
    if request.method == 'POST':
        # Log activity
        ActivityLog.objects.create(
            project=project,
            user=request.user,
            action='photo_deleted',
            description=f"Deleted photo: {photo.title or 'Untitled'}"
        )
        
        photo.delete()
        messages.success(request, 'Photo deleted successfully!')
        return redirect('project_photos', project_pk=project.pk)
    
    return render(request, 'tracker/photo_confirm_delete.html', {
        'photo': photo,
        'project': project
    })


# ==================== Activity Timeline ====================

@login_required
def project_timeline(request, pk):
    """View project activity timeline."""
    project = get_object_or_404(Project, pk=pk)
    activities = project.activity_logs.all()[:50]  # Last 50 activities
    
    context = {
        'project': project,
        'activities': activities,
    }
    return render(request, 'tracker/project_timeline.html', context)


# ==================== Budget Alerts ====================

@login_required
def budget_alerts(request):
    """View all budget alerts for user's projects."""
    user_projects = Project.objects.filter(created_by=request.user)
    alerts = BudgetAlert.objects.filter(
        project__in=user_projects
    ).order_by('-created_at')[:20]
    
    unread_count = alerts.filter(is_read=False).count()
    
    context = {
        'alerts': alerts,
        'unread_count': unread_count,
    }
    return render(request, 'tracker/budget_alerts.html', context)


@login_required
def mark_alert_read(request, pk):
    """Mark an alert as read."""
    alert = get_object_or_404(BudgetAlert, pk=pk)
    alert.is_read = True
    alert.save()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    return redirect('budget_alerts')


def check_budget_alerts(project):
    """Helper function to check and create budget alerts."""
    percentage = project.budget_utilization_percentage
    
    # Check for different alert levels
    if percentage >= 100 and not project.budget_alerts.filter(alert_type='exceeded').exists():
        BudgetAlert.objects.create(
            project=project,
            alert_type='exceeded',
            percentage=percentage,
            message=f"Budget exceeded! Spent Ksh{project.total_spent} of Ksh{project.budget} ({percentage:.1f}%)"
        )
    elif percentage >= 90 and not project.budget_alerts.filter(alert_type='critical').exists():
        BudgetAlert.objects.create(
            project=project,
            alert_type='critical',
            percentage=percentage,
            message=f"Critical: {percentage:.1f}% of budget used (Ksh{project.total_spent} of Ksh{project.budget})"
        )
    elif percentage >= 75 and not project.budget_alerts.filter(alert_type='warning').exists():
        BudgetAlert.objects.create(
            project=project,
            alert_type='warning',
            percentage=percentage,
            message=f"Warning: {percentage:.1f}% of budget used (Ksh{project.total_spent} of Ksh{project.budget})"
        )


